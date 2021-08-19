import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook

class CustomerOrder():
    customerList = []

    def __init__(self, customerName) -> None:
        self.customer = customerName
        self.customerList.append(customerName)
        self.order = {}
        self.totalItemQty = 0

    def addOrder(self, itemCode, itemQtyInCtn):
        self.order[itemCode] = itemQtyInCtn
        self.totalItemQty += int(itemQtyInCtn)
        return(self.order)

    def getQtyByItemCode(self, itemCode):
        return(self.order[itemCode])

    def getOrder(self):
        return(self.order)
    
    def getCustomer(self):
        return(self.customer)


def importData(fileList):
    output = []
    for file in fileList:
        newOrder = None
        countBreak = 0
        print(f'Openning file: {file}')
        wb = load_workbook(filename = importPath + '\\' + file)
        wsheet = wb.active
        customerName = wsheet['C9'].value
        newOrder = CustomerOrder(customerName)
        for row in range(1, wsheet.max_row):
            itemCode = str(wsheet.cell(row, 2).value).strip('="')
            itemQty = str(wsheet.cell(row, 5).value)
            if re.match(r'CHB-', itemCode):
                #print(itemCode, itemQty, countBreak)
                if itemQty.isnumeric():
                    #print(customerName + ': ' + itemCode + ' - ' + itemQty)
                    newOrder.addOrder(itemCode, itemQty)       
            if itemCode is None:
                countBreak += 1
            if countBreak > 20:
                break
        output.append(newOrder)
    return(output)


def importTemplate(templateFile):
    wbTemplate = load_workbook(templateFile)
    wsheetTemplate = wbTemplate.active
    countBreak = 0
    output = []
    for row in range(1, wsheetTemplate.max_row+1):
        item = str(wsheetTemplate.cell(row, 1).value)
        output.append(item)
        if item is None:
            countBreak += 1
        if countBreak > 20:
            break
    return(output)


def generateHeader(data):
    output = ['']
    for item in data:
        output.append(item.customer)
    return(output)
    #outputData.append(placeholder0)


def generateData(data, template):
    output = []
    for itemCode in template:
        placeholder = []
        placeholder.append(itemCode)
        if re.match(r'CHB-', itemCode):
            for order in data:
                itemQty = order.order.get(itemCode, '')
                if itemQty != '':
                    placeholder.append(int(itemQty))
                else:
                    placeholder.append('')
        output.append(placeholder)
    return(output)


def generateSUM(data):
    output = ['SUM']
    for n in range(len(data[0])-1):
        output.append(0)
    for row in data:
        for col in range(1, len(row)):
            if row[col] != '' and str(row[col]).isnumeric():
                output[col] += int(row[col])
    return(output)
                

def generateTotalOrder(data):
    output = ['Orders from Orderlist']
    for order in data:
        output.append(order.totalItemQty)
    return(output)

def dataValidation(rawData, outputData, templateList):
    sum = {}
    print('\n-------------------------------')
    print("Data Validation Result")
    print('-------------------------------')
    for row in outputData:
        if row[0] == 'SUM':
            for item in range(1, len(row)):
                sum[outputData[0][item]] = row[item]
    for customer in rawData:
        if customer.totalItemQty != sum[customer.customer]:
            print(f'<{customer.customer}>:')
            print(f'Total from Order Form: {customer.totalItemQty}')
            print(f'Total from output: {sum[customer.customer]}\n')
            for itemCode, itemQty in customer.order.items():
                if itemCode not in templateList:
                    print(f'Item code {itemCode} - Qty {itemQty} is missing from the template.')
        else:
            print(f'<{customer.customer}> - OK!')
            print(f'Total from Order Form: {customer.totalItemQty}')
            print(f'Total from output: {sum[customer.customer]}\n')
        print('-------------------------------')
    print('\n')
    

def output2xlsx(output):
    wbResult = Workbook()
    resultFilename = outputPath + datetime.now().strftime('%d-%m-%Y_%H%M%S.xlsx')
    wsResult = wbResult.active
    for row in output:
        wsResult.append(row)
    wbResult.save(resultFilename)


# Fix Variable
dir_path = os.path.dirname(os.path.realpath(__file__))
importPath = dir_path + '/Input/'
outputPath = dir_path + '/Output/'
templateFile = dir_path + '/template.xlsx'


# Initialize
rawData = []
templateData = []
outputData = []
importFiles = os.listdir(importPath)

# Main
rawData = importData(importFiles)
templateData = importTemplate(templateFile)
outputData.append(generateHeader(rawData))
outputData += generateData(rawData, templateData)
outputData.append(generateSUM(outputData))
outputData.append(generateTotalOrder(rawData))
dataValidation(rawData, outputData, templateData)
output2xlsx(outputData)
print('\n')
input('Press ENTER to continue...')
