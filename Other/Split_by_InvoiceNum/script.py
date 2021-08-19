import os, csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins



# Config
dir_path = os.path.dirname(os.path.realpath(__file__))
input_dir = dir_path + "/input/"
output_dir = dir_path + "/output/"
width_file = dir_path + "/config/width.csv"


# Functions
def getUniqueInvoiceNo(ws):
    invoiceList = []
    for i in range(2, ws.max_row):
        invoiceNum = ws.cell(column=29, row=i).value
        if invoiceNum not in invoiceList:
            invoiceList.append(invoiceNum)
    #print(invoiceList)
    return(invoiceList)       

def getUniqueInvoiceNoCSV(dest_filename):
    invoiceList = []
    file = open(dest_filename)
    csvfile = csv.reader(file)
    for line in csvfile:
        invoiceNum = line[28]
        if invoiceNum not in invoiceList:
            invoiceList.append(invoiceNum)
    invoiceList.pop(0)
    print(invoiceList)
    return(invoiceList)       

def categorizebyInvoiceNum(ws, invoiceList):
    outputAllFiles = []
    for invoiceNum in invoiceList:
        outputSingleFile = {}
        outputSingleFile['invoiceNum'] = invoiceNum
        fileOutput = []
        for i in range(1, ws.max_row):
            checkInvoiceNum = ws.cell(column=29, row=i).value
            if i == 1 or invoiceNum == checkInvoiceNum:
                fileOutput.append(getRowToList(ws,i))
        outputSingleFile['output'] = fileOutput
        outputSingleFile_copy = outputSingleFile.copy()
        outputAllFiles.append(outputSingleFile_copy)
        #print(outputSingleFile)
    return(outputAllFiles)

def categorizebyInvoiceNumCSV(dest_filename, invoiceList):
    outputAllFiles = []
    for invoiceNum in invoiceList:
        outputSingleFile = {}
        outputSingleFile['invoiceNum'] = invoiceNum
        fileOutput = []
        i = 0
        file = open(dest_filename)
        csvfile = csv.reader(file)
        for line in csvfile:
            checkInvoiceNum = line[28]
            if i == 0 or invoiceNum == checkInvoiceNum:
                fileOutput.append(line)
                i += 1
        outputSingleFile['output'] = fileOutput
        outputSingleFile_copy = outputSingleFile.copy()
        outputAllFiles.append(outputSingleFile_copy)
    return(outputAllFiles)

def getColWidth(ws):
    colWidthList = []
    for letter in range(1,ws.max_column):
        colWidth = ws.column_dimensions[get_column_letter(letter)].width
        colWidthList.append(colWidth)
    return(colWidthList)

def getRowToList(ws, rowCount):
    rowData = []
    for row in ws[rowCount]:
        rowData.append(row.value)
    return(rowData)

def outputfilestoxlsx(outputFiles):
    for file in outputFiles:
        filename = output_dir + file['invoiceNum'] + '.xlsx'
        wb2print = Workbook()
        ws2print = wb2print.active
        for row in file['output']:
            ws2print.append(row)
        applyDateFormat(ws2print)
        applyNumFormat(ws2print)
        applyTextWrap(ws2print)
        applyRowHeight(ws2print)
        applyColWidth(ws2print)
        hideColumn(ws2print)
        applyPageFormat(ws2print)
        wb2print.save(filename)
        print('-->> ' + filename)

def applyDateFormat(ws):
    col = ws['B']
    for cell in col:
        cell.number_format = 'mmm-yy'

def applyNumFormat(ws):
    col = ws['AM']
    for cell in col:
        cell.number_format = '0'

def applyTextWrap(ws):
    row = ws[1]
    for cell in row:
        cell.alignment = Alignment(wrap_text=True,vertical='bottom')

def applyRowHeight(ws):
    ws.row_dimensions[1].height = 109.5

def applyColWidth(ws):
    startCol = 1
    file = open(width_file)
    csvfile = csv.reader(file)
    #for letter in range(1,ws.max_column):
    for width in next(csvfile):
        ws.column_dimensions[get_column_letter(startCol)].width = width
        startCol += 1

def hideColumn(ws):
    col2hide = ['F','G','H','I','J','K','L','AE','AF','AG','AH','AI','AJ','AK','AN','AO']
    for col in col2hide:
        ws.column_dimensions[col].hidden= True

def applyPageFormat(ws):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize = 9
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0.2, footer=0.2)


# Main
input_filenames = os.listdir(input_dir)
print('===================================================================================================')
print('-------------------------------------<< BEGIN TO RUN SCRIPT>>--------------------------------------')
print('===================================================================================================')

for filename in input_filenames:
    dest_filename = input_dir + filename

    # Excel Format
    #wb = load_workbook(filename = dest_filename, data_only=True)
    #ws = wb.active
    #uniqueInvoiceList = getUniqueInvoiceNo(ws)
    #outputFilesData = categorizebyInvoiceNum(ws, uniqueInvoiceList)
    #colWidthList = getColWidth(ws)

    # CSV 
    print('Processing file: ' + filename)
    print('----------------------------------------------------------------------------------------')
    print('List of Invoice Number in this file: ')
    uniqueInvoiceList = getUniqueInvoiceNoCSV(dest_filename)
    print('\nOutputing files: ')
    outputFilesData = categorizebyInvoiceNumCSV(dest_filename, uniqueInvoiceList)
    outputfilestoxlsx(outputFilesData)
    print('===================================================================================================')

print('---------------------------------<< END SUCCESSFULLY COMPLETED >>----------------------------------')
print('===================================================================================================')
print('\n')
input('Press ENTER to continue...')
